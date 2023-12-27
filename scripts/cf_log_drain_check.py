# Pretty quick script/copied and pasted from others to meet the requirement of determining which apps are bound to a log drain.
# Ended up loading all service instances and creditionals at once as it seemed faster than calling the API mutliple times (per space,app).
# Didn't implement filtering on orgs, spaces, apps but it should be easy to do so if it becomes a requirement.
# An additional check could be added to show log drains not bound to an app but that wasn't my initial scope.

import sys
import subprocess
import os
import json
import hashlib
import getopt
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


######################################################################################################
# Globals setup and directory location initiation
######################################################################################################

per_page = 5000
end_point_version = "/v3/"
reporting_object_array = []
log_drain_object_array = []

######################################################################################################
# Generic command line functions.
######################################################################################################


def get_cf_response(url):
    if os.name == "nt":
        # Need to escape the URL ampersands when running on Windows
        url = url.replace("&", "^&")
        p = subprocess.Popen(
            ["cf.exe", "curl", url], stdout=subprocess.PIPE, shell=True
        )
    else:
        p = subprocess.Popen(["cf", "curl", url], stdout=subprocess.PIPE)

    if p.returncode == 1:
        print("Can't find the CF CLI executable, is it installed?")
        os.system.exit(1)

    output = p.communicate()[0]
    data = json.loads(output.decode("utf-8"))
    return data


def run_api_cmd_resources(url):
    all_data = {"resources": []}
    next_page = True

    while next_page:
        data = get_cf_response(url)

        if data["pagination"]["total_results"] == 0:
            next_page = False
        else:
            for resource in data["resources"]:
                all_data["resources"].append(resource)

                try:
                    # if this element exists, a subsequent page exists so get the new URL and continue iteration
                    url = data["pagination"]["next"]["href"]
                    index = url.find(end_point_version)
                    url = url[index:]
                except:
                    # break out of loop because this was the last page
                    next_page = False

    return all_data


######################################################################################################
######################################################################################################
def obtain_cf_service_data(
    org_guid, org_name, space_name, space_guid, app_guid, app_name
):
    appjson = get_cf_response("/v3/apps/" + app_guid)

    for d in appjson["resources"]:
        print(
            org_name
            + ","
            + space_name
            + ","
            + app_name
            + ","
            + str(d["instances"])
            + ","
            + str(d["memory_in_mb"])
            + ","
            + str(d["disk_in_mb"])
        )
    return


######################################################################################################
######################################################################################################


######################################################################################################
######################################################################################################
def obtain_cf_data_for_entire_space(org_guid, org_name, space_guid, space_name):
    appsjson = run_api_cmd_resources(
        "/v3/apps?order_by=name&per_page="
        + str(per_page)
        + "&space_guids="
        + space_guid
    )

    print("Obtaining application list for " + space_name)
    for d in appsjson["resources"]:
        rep_object_entry = {
            "org_guid": org_guid,
            "org_name": org_name,
            "space_guid": space_guid,
            "space_name": space_name,
            "app_guid": d["guid"],
            "app_name": d["name"],
            "log_drain_name": "",
            "log_drain_guid": "",
            "log_drain_created_date": "",
            "log_drain_updated_date": "",
            "log_drain_bind_guid": "",
            "log_drain_bind_created_date": "",
            "log_drain_bind_updated_date": "",
        }

        reporting_object_array.append(rep_object_entry)

    return


######################################################################################################
######################################################################################################


######################################################################################################
######################################################################################################
def obtain_cf_data_for_entire_org(org_guid, org_name):
    spacesjson = run_api_cmd_resources(
        "/v3/spaces?order_by=name&per_page="
        + str(per_page)
        + "&organization_guids="
        + org_guid
    )

    print("Obtaining space data for " + org_name)
    for d in spacesjson["resources"]:
        obtain_cf_data_for_entire_space(org_guid, org_name, d["guid"], d["name"])

    return


######################################################################################################
######################################################################################################


######################################################################################################
######################################################################################################
def obtain_log_drain_service_information(log_drain_url):
    print("Obtaining available log drain service listings")
    service_instances = run_api_cmd_resources(
        "/v3/service_instances?per_page=" + str(per_page)
    )

    print("Filtering for Log Drain services")
    temp_log_drain_object_array = []
    for d in service_instances["resources"]:
        if (
            d.get("syslog_drain_url") is not None
            and d["syslog_drain_url"] != ""
            and d["syslog_drain_url"].lower() == log_drain_url
        ):
            temp_log_drain_object_array.append(
                {
                    "service_instance_name": d["name"],
                    "service_instance_guid": d["guid"],
                    "service_instance_created_date": d["created_at"],
                    "service_instance_updated_date": d["updated_at"],
                }
            )

    print("Obtaining log drain to application bindings")

    temp_service_creds = run_api_cmd_resources(
        "/v3/service_credential_bindings?per_page=" + str(per_page)
    )
    for d in temp_service_creds["resources"]:
        for temp_obj in temp_log_drain_object_array:
            if (
                d["relationships"]["service_instance"]["data"]["guid"]
                == temp_obj["service_instance_guid"]
            ):
                log_drain_object_array.append(
                    {
                        "service_instance_name": temp_obj["service_instance_name"],
                        "service_instance_guid": temp_obj["service_instance_guid"],
                        "service_instance_created_date": temp_obj[
                            "service_instance_created_date"
                        ],
                        "service_instance_updated_date": temp_obj[
                            "service_instance_updated_date"
                        ],
                        "service_credential_bindings_created_date": d["created_at"],
                        "service_credential_bindings_updated_date": d["updated_at"],
                        "service_credential_bindings_guid": d["guid"],
                        "app_guid": d["relationships"]["app"]["data"]["guid"],
                    }
                )

    return


######################################################################################################
######################################################################################################
def combine_log_drain_and_app_data():
    print("Combining log drain and application data")

    for drain_obj in log_drain_object_array:
        for app_obj in reporting_object_array:
            if drain_obj["app_guid"] == app_obj["app_guid"]:
                app_obj["log_drain_name"] = drain_obj["service_instance_name"]
                app_obj["log_drain_guid"] = drain_obj["service_instance_guid"]
                app_obj["log_drain_created_date"] = drain_obj[
                    "service_instance_created_date"
                ]
                app_obj["log_drain_updated_date"] = drain_obj[
                    "service_instance_updated_date"
                ]
                app_obj["log_drain_bind_created_date"] = drain_obj[
                    "service_credential_bindings_created_date"
                ]
                app_obj["log_drain_bind_updated_date"] = drain_obj[
                    "service_credential_bindings_updated_date"
                ]
                app_obj["log_drain_bind_guid"] = drain_obj[
                    "service_credential_bindings_guid"
                ]

    return


######################################################################################################
######################################################################################################
def export_results(output_file):
    print("Exporting results")

    Excelworkbook = Workbook()
    audit_sheet = Excelworkbook.create_sheet("Log Drain Audit", 1)
    Excelworkbook.active = 1

    HeaderColorFill = PatternFill(fgColor="C0C0C0", fill_type="solid")

    audit_sheet["A1"] = "Organization Name"
    audit_sheet["A1"].fill = HeaderColorFill
    audit_sheet["A1"].font = Font(bold=True)

    audit_sheet["B1"] = "Space Name"
    audit_sheet["B1"].fill = HeaderColorFill
    audit_sheet["B1"].font = Font(bold=True)

    audit_sheet["C1"] = "Application Name"
    audit_sheet["C1"].fill = HeaderColorFill
    audit_sheet["C1"].font = Font(bold=True)

    audit_sheet["D1"] = "Log Drain Name"
    audit_sheet["D1"].fill = HeaderColorFill
    audit_sheet["D1"].font = Font(bold=True)

    audit_sheet["E1"] = "Log Drain GUID"
    audit_sheet["E1"].fill = HeaderColorFill
    audit_sheet["E1"].font = Font(bold=True)

    audit_sheet["F1"] = "Log Drain Created Date"
    audit_sheet["F1"].fill = HeaderColorFill
    audit_sheet["F1"].font = Font(bold=True)

    audit_sheet["G1"] = "Log Drain Updated Date"
    audit_sheet["G1"].fill = HeaderColorFill
    audit_sheet["G1"].font = Font(bold=True)

    audit_sheet["H1"] = "Log Drain Binding GUID"
    audit_sheet["H1"].fill = HeaderColorFill
    audit_sheet["H1"].font = Font(bold=True)

    audit_sheet["I1"] = "Log Drain Binding Created Date"
    audit_sheet["I1"].fill = HeaderColorFill
    audit_sheet["I1"].font = Font(bold=True)

    audit_sheet["J1"] = "Log Drain Binding Updated Date"
    audit_sheet["J1"].fill = HeaderColorFill
    audit_sheet["J1"].font = Font(bold=True)

    ExcelRowRef = 1
    for obj in reporting_object_array:
        ExcelRowRef += 1
        audit_sheet["A" + str(ExcelRowRef)] = obj["org_name"]
        audit_sheet["B" + str(ExcelRowRef)] = obj["space_name"]
        audit_sheet["C" + str(ExcelRowRef)] = obj["app_name"]
        if obj["log_drain_guid"] == "":
            audit_sheet["C" + str(ExcelRowRef)].font = Font(color="FF0000", bold=True)
        audit_sheet["D" + str(ExcelRowRef)] = obj["log_drain_name"]
        audit_sheet["E" + str(ExcelRowRef)] = obj["log_drain_guid"]
        audit_sheet["F" + str(ExcelRowRef)] = obj["log_drain_created_date"]
        audit_sheet["G" + str(ExcelRowRef)] = obj["log_drain_updated_date"]
        audit_sheet["H" + str(ExcelRowRef)] = obj["log_drain_bind_guid"]
        audit_sheet["I" + str(ExcelRowRef)] = obj["log_drain_bind_created_date"]
        audit_sheet["J" + str(ExcelRowRef)] = obj["log_drain_bind_updated_date"]

    for col in audit_sheet.columns:
        max_lenght = 0
        col_name = re.findall("\w\d", str(col[0]))
        col_name = col_name[0]
        col_name = re.findall("\w", str(col_name))[0]
        for cell in col:
            try:
                if len(str(cell.value)) > max_lenght:
                    max_lenght = len(cell.value)
            except:
                pass
        adjusted_width = max_lenght + 2
        audit_sheet.column_dimensions[col_name].width = adjusted_width

    Excelworkbook.save(filename=output_file)

    return


######################################################################################################
######################################################################################################
######################################################################################################
# Starting point for script logic
######################################################################################################
print("")

full_cmd_arguments = sys.argv
argument_list = full_cmd_arguments[1:]
short_options = "l:f:"
long_options = ["log_drain_url=", "report_file="]

try:
    arguments, values = getopt.getopt(argument_list, short_options, long_options)
except getopt.error as err:
    print(str(err))
    sys.exit(2)

log_drain_url = ""
output_file = ""
for current_argument, current_value in arguments:
    if current_argument in ("-f", "--report_file"):
        output_file = current_value
    if current_argument in ("-l", "--log_drain_url"):
        log_drain_url = current_value

if log_drain_url == "":
    print("A log drain URL must be specified.")
    sys.exit(-1)

if output_file == "":
    print("An MS Excel based output file (i.e., .xlsx) must be specified.")
    sys.exit(-2)

orgjson = run_api_cmd_resources(
    "/v3/organizations?order_by=name&per_page=" + str(per_page) + ""
)

for d in orgjson["resources"]:
    obtain_cf_data_for_entire_org(d["guid"], d["name"])

obtain_log_drain_service_information(log_drain_url)

combine_log_drain_and_app_data()

export_results(output_file)
