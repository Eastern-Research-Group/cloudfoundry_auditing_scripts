######################################################################################################
# Background docs: https://cloud.gov/docs/compliance/auditing-activity/.
# There is a ton of duplicated code in here that I plan to refactor. Basically, v3 of the API
# seems much more consistent than v2 which will allow us to consolidate code sections.
######################################################################################################

import subprocess
import sys
import getopt
import os
import json
import hashlib
import re
import datetime
import dateutil.parser
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

start_path = os.getcwd()
data_folder_path = start_path
cf_guid_dict = dict()
per_page = 5000

######################################################################################################
# Generic command line functions.
######################################################################################################


def run_cmd_suppress_output(cmd, args):
    subprocess.call([cmd, args], stdout=open(
        os.devnull, "w"), stderr=subprocess.STDOUT)
    return


def run_cli_cmd(cmd, filename):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    f = open(filename + ".txt", "w")
    print(output.decode(), file=f)
    f.close()
    return


def run_api_cmd(cmd, filename):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    f = open(filename + ".json", "w")
    print(output.decode(), file=f)
    f.close()
    return


def run_api_cmd_rtn_json(cmd):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    return output.decode()


def run_api_cmd_and_hash_output_txt(cmd, hash_statement, json_attrib, filename):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    json_data = json.loads(output)

    text = ""
    if "CF-NotAuthorized" in output.decode():
        text = "Not authorized to perform action around this action."
        hash_statement = text
        print(text + " " + '/'.join(cmd))
    elif json_attrib != None and json_attrib not in output.decode():
        text = "There is no " + json_attrib + " in the results of this action."
        hash_statement = text
    elif json_attrib == None:
        text = json_data
    else:
        text = json_data[json_attrib]

    hash_object = hashlib.md5(json.dumps(text).encode('utf-8'))
    f = open(filename + ".txt", "w")
    print(hash_statement + ": " + hash_object.hexdigest(), file=f)
    f.close()
    return


def versiontuple(v):
    return tuple(map(int, (v.split("."))))


def remove_prefix(s, prefix):
    return s[len(prefix):] if s.startswith(prefix) else s


def GetKey(val):
    for key, value in cf_guid_dict.items():
        if val == value:
            return key
    return None

######################################################################################################
######################################################################################################


def buildpack_by_app_check(org_query, space_query, b_audit_sheet, buildpacks_json, cf_api_endpoint):

    HeaderColorFill = PatternFill(fgColor="C0C0C0", fill_type="solid")

    b_audit_sheet["A1"] = "Organization Name"
    b_audit_sheet["A1"].fill = HeaderColorFill
    b_audit_sheet["A1"].font = Font(bold=True)

    b_audit_sheet["B1"] = "Space Name"
    b_audit_sheet["B1"].fill = HeaderColorFill
    b_audit_sheet["B1"].font = Font(bold=True)

    b_audit_sheet["C1"] = "Application Name"
    b_audit_sheet["C1"].fill = HeaderColorFill
    b_audit_sheet["C1"].font = Font(bold=True)

    b_audit_sheet["D1"] = "Application State"
    b_audit_sheet["D1"].fill = HeaderColorFill
    b_audit_sheet["D1"].font = Font(bold=True)

    b_audit_sheet["E1"] = "Buildpack Name"
    b_audit_sheet["E1"].fill = HeaderColorFill
    b_audit_sheet["E1"].font = Font(bold=True)

    b_audit_sheet["F1"] = "Buildpack Status"
    b_audit_sheet["F1"].fill = HeaderColorFill
    b_audit_sheet["F1"].font = Font(bold=True)

    b_audit_sheet["G1"] = "Application Buildpack Version"
    b_audit_sheet["G1"].fill = HeaderColorFill
    b_audit_sheet["G1"].font = Font(bold=True)

    b_audit_sheet["H1"] = "Available Buildpack Version"
    b_audit_sheet["H1"].fill = HeaderColorFill
    b_audit_sheet["H1"].font = Font(bold=True)

    b_audit_sheet["I1"] = "Last Application Updated On"
    b_audit_sheet["I1"].fill = HeaderColorFill
    b_audit_sheet["I1"].font = Font(bold=True)

    b_audit_sheet["J1"] = "Available Buildpack Updated At"
    b_audit_sheet["J1"].fill = HeaderColorFill
    b_audit_sheet["J1"].font = Font(bold=True)

    b_audit_sheet["K1"] = "App Link"
    b_audit_sheet["K1"].fill = HeaderColorFill
    b_audit_sheet["K1"].font = Font(bold=True)

    temp = run_api_cmd_rtn_json(
        ["cf", "curl", "/v3/apps?order_by=name&per_page=" + str(per_page) + ""+org_query+space_query])
    done = False
    ExcelRowRef = 1
    while done != True:
        data = json.loads(temp)
        for d in data["resources"]:
            number_of_app_buildpacks = 0
            if d["state"] == "STOPPED":
                number_of_app_buildpacks = len(
                    d["lifecycle"]["data"]["buildpacks"])
            else:
                droplet_txt = run_api_cmd_rtn_json(["cf", "curl", remove_prefix(
                    d["links"]["current_droplet"]["href"], cf_api_endpoint)])
                droplet_json = json.loads(droplet_txt)
                if "buildpacks" in droplet_json:
                    number_of_app_buildpacks = len(droplet_json["buildpacks"])
                else:
                    print("Error getting buildpack information: " +
                          json.dumps(droplet_json))

            org_txt = run_api_cmd_rtn_json(["cf", "curl", remove_prefix(
                d["links"]["space"]["href"], cf_api_endpoint)])
            org_json = json.loads(org_txt)
            org = cf_guid_dict.get(
                org_json["relationships"]["organization"]["data"]["guid"], "UNKNOWN ORG")
            space_name = cf_guid_dict.get(
                d["relationships"]["space"]["data"]["guid"], "UNKNOWN SPACE")
            if org != "UNKNOWN ORG" and space_name != "UNKNOWN SPACE":
                space_name = remove_prefix(space_name, org+"-")

            for x in range(0, number_of_app_buildpacks):
                print(org + ", " + space_name + ", " + d["name"])
                outdated_buildpack = False
                ExcelRowRef += 1
                b_audit_sheet["A" + str(ExcelRowRef)] = org
                b_audit_sheet["B" + str(ExcelRowRef)] = space_name
                b_audit_sheet["C" + str(ExcelRowRef)] = d["name"]
                b_audit_sheet["D" + str(ExcelRowRef)] = d["state"]

                buildpack_name_from_json = ""
                if d["state"] == "STOPPED":
                    buildpack_name_from_json = d["lifecycle"]["data"]["buildpacks"][x]
                else:
                    buildpack_name_from_json = droplet_json["buildpacks"][x]["name"]

                b_audit_sheet["E" + str(ExcelRowRef)
                              ] = buildpack_name_from_json

                b_ver = ""
                b_update_date = ""
                for bpack_2 in buildpacks_json["resources"]:
                    if bpack_2["name"] == buildpack_name_from_json:
                        b_ver = bpack_2["version"][1:]
                        b_update_date = bpack_2["updated_at"]

                    if d["state"] == "STOPPED":
                        b_audit_sheet["G" + str(ExcelRowRef)] = "Not available"
                        b_audit_sheet["G" + str(ExcelRowRef)
                                      ].font = Font(color='FF0000', bold=True)
                        outdated_buildpack = True
                    else:
                        if droplet_json["buildpacks"][x]["version"] is None:
                            b_audit_sheet["G" +
                                          str(ExcelRowRef)] = "Not available"
                        else:
                            b_audit_sheet["G" + str(ExcelRowRef)
                                          ] = droplet_json["buildpacks"][x]["version"]
                        if b_ver is not None and b_ver != "" and droplet_json["buildpacks"][x]["version"] is not None and droplet_json["buildpacks"][x]["version"] != "":
                            if versiontuple(b_ver) > versiontuple(droplet_json["buildpacks"][x]["version"]):
                                b_audit_sheet["G" + str(ExcelRowRef)
                                              ].font = Font(color='FF0000', bold=True)
                                outdated_buildpack = True

                b_audit_sheet["H" + str(ExcelRowRef)] = b_ver
                b_audit_sheet["I" + str(ExcelRowRef)] = d["updated_at"]

                if d["updated_at"] is not None and d["updated_at"] != "" and b_update_date is not None and b_update_date != "":
                    if dateutil.parser.parse(d["updated_at"]) < dateutil.parser.parse(b_update_date):
                        b_audit_sheet["J" + str(ExcelRowRef)
                                      ].font = Font(color='FF0000', bold=True)
                        outdated_buildpack = True

                if outdated_buildpack == True:
                    b_audit_sheet["F" + str(ExcelRowRef)] = "Outdated"
                    b_audit_sheet["F" + str(ExcelRowRef)
                                  ].font = Font(color='FF0000', bold=True)
                else:
                    b_audit_sheet["F" + str(ExcelRowRef)] = "Current"

                b_audit_sheet["J" + str(ExcelRowRef)] = b_update_date
                b_audit_sheet["K" + str(ExcelRowRef)] = "cf curl " + remove_prefix(
                    d["links"]["self"]["href"], cf_api_endpoint)

        if data["pagination"]["next"] is not None:
            href_link = data["pagination"]["next"]["href"]
        else:
            href_link = "null"

        if href_link != "null":
            href_link = remove_prefix(href_link, cf_api_endpoint)
            temp = run_api_cmd_rtn_json(["cf", "curl", href_link])
        else:
            done = True

    for col in b_audit_sheet.columns:
        max_lenght = 0
        col_name = re.findall('\w\d', str(col[0]))
        col_name = col_name[0]
        col_name = re.findall('\w', str(col_name))[0]
        for cell in col:
            try:
                if len(str(cell.value)) > max_lenght:
                    max_lenght = len(cell.value)
            except:
                pass
        adjusted_width = (max_lenght+2)
        b_audit_sheet.column_dimensions[col_name].width = adjusted_width

    return


######################################################################################################
######################################################################################################

def buildpack_audit(org_name, space_names, b_audit_sheet, buildpacks_json, cf_api_endpoint):
    print("==================================================================")
    print("Recording buildpack by app information to workbook...")
    print("==================================================================")
    orgs_json = run_api_cmd_rtn_json(
        ["cf", "curl", "/v3/organizations?order_by=name&per_page=" + str(per_page) + ""])
    data = json.loads(orgs_json)

    for d in data["resources"]:
        cf_guid_dict.update({d["guid"]: d["name"]})
        spaces_json = run_api_cmd_rtn_json(
            ["cf", "curl", "/v3/spaces?order_by=name&per_page=" + str(per_page) + "&organization_guids="+d["guid"]])
        spaces_data = json.loads(spaces_json)
        for s in spaces_data["resources"]:
            cf_guid_dict.update({s["guid"]: d["name"]+"-"+s["name"]})

    org_query = ""
    org_guid = None
    if org_name != "":
        org_guid = GetKey(org_name)
        if org_guid == None:
            print("Could not find organization GUID for ORG NAME: " + org_name)
            sys.exit(3)
        else:
            org_query = "&organization_guids=" + org_guid

    space_query = ""
    space_guids = []
    if space_names != "":
        for value in space_names.split(","):
            space_guid = GetKey(org_name + "-" + value)
            if space_guid == None:
                print("Could not find space GUID for SPACE NAME: " + value)
                sys.exit(4)
            else:
                space_guids.append(space_guid)
        space_query = "&space_guids=" + ",".join(space_guids)

    buildpack_by_app_check(org_query, space_query,
                           b_audit_sheet, buildpacks_json, cf_api_endpoint)

    return

######################################################################################################
######################################################################################################


def buildpack_workbook(Excelworkbook, buildpacks_json, cf_api_endpoint):
    print("==================================================================")
    print("Recording buildpack information to workbook...")
    print("==================================================================")
    os.chdir(data_folder_path)
    HeaderColorFill = PatternFill(fgColor="C0C0C0", fill_type="solid")

    sheet0 = Excelworkbook.active
    sheet0.title = "Available Buidpacks"

    sheet0["A1"] = "Buildpack Name"
    sheet0["A1"].fill = HeaderColorFill
    sheet0["A1"].font = Font(bold=True)

    sheet0["B1"] = "Buildpack Version"
    sheet0["B1"].fill = HeaderColorFill
    sheet0["B1"].font = Font(bold=True)

    sheet0["C1"] = "Last Updated On"
    sheet0["C1"].fill = HeaderColorFill
    sheet0["C1"].font = Font(bold=True)

    sheet0["D1"] = "Buildpack Stack"
    sheet0["D1"].fill = HeaderColorFill
    sheet0["D1"].font = Font(bold=True)

    sheet0["E1"] = "Buildpack GUID"
    sheet0["E1"].fill = HeaderColorFill
    sheet0["E1"].font = Font(bold=True)

    sheet0["F1"] = "Buildpack Link"
    sheet0["F1"].fill = HeaderColorFill
    sheet0["F1"].font = Font(bold=True)

    ExcelRowRef = 1
    for d in buildpacks_json["resources"]:
        ExcelRowRef += 1
        sheet0["A" + str(ExcelRowRef)] = d["name"]
        sheet0["B" + str(ExcelRowRef)] = d["version"]
        sheet0["C" + str(ExcelRowRef)] = d["updated_at"]
        sheet0["D" + str(ExcelRowRef)] = d["stack"]
        sheet0["E" + str(ExcelRowRef)] = d["guid"]
        sheet0["F" + str(ExcelRowRef)] = "cf curl " + remove_prefix(
            d["links"]["self"]["href"], cf_api_endpoint)

    for col in sheet0.columns:
        max_lenght = 0
        col_name = re.findall('\w\d', str(col[0]))
        col_name = col_name[0]
        col_name = re.findall('\w', str(col_name))[0]
        for cell in col:
            try:
                if len(str(cell.value)) > max_lenght:
                    max_lenght = len(cell.value)
            except:
                pass
        adjusted_width = (max_lenght+2)
        sheet0.column_dimensions[col_name].width = adjusted_width

    return
######################################################################################################
######################################################################################################


def start(org_name, space_names, output_file, cf_api_endpoint):
    buildpacks_text = run_api_cmd_rtn_json(
        ["cf", "curl", "/v3/buildpacks?per_page=" + str(per_page) + ""])
    buildpacks_json = json.loads(buildpacks_text)

    for d in buildpacks_json["resources"]:
        version = d["filename"][len(
            d["name"] + "-"+d["stack"]+"-"):len(d["filename"])]
        version = version.replace(".zip", "")
        d["version"] = version

    Excelworkbook = Workbook()
    buildpack_workbook(Excelworkbook, buildpacks_json, cf_api_endpoint)
    b_audit_sheet = Excelworkbook.create_sheet("Buildpack Audit", 1)
    Excelworkbook.active = 1
    buildpack_audit(org_name, space_names, b_audit_sheet,
                    buildpacks_json, cf_api_endpoint)

    Excelworkbook.save(filename=output_file)

    return
######################################################################################################
######################################################################################################


full_cmd_arguments = sys.argv
argument_list = full_cmd_arguments[1:]
short_options = "o:s:f:c:"
long_options = ["organization=", "space=", "report_file=", "cf_api_endpoint="]
output_file = ""
cf_api_endpoint = ""


try:
    arguments, values = getopt.getopt(
        argument_list, short_options, long_options)
except getopt.error as err:
    print(str(err))
    sys.exit(2)

o_name = ""
s_names = ""
for current_argument, current_value in arguments:
    if current_argument in ("-o", "--organization"):
        o_name = current_value
    if current_argument in ("-s", "--space"):
        s_names = current_value
    if current_argument in ("-f", "--report_file"):
        output_file = current_value
    if current_argument in ("-c", "--cf_api_endpoint"):
        cf_api_endpoint = current_value

if o_name == "" and s_names != "":
    print("Cannot specify spaces without specifying a organization.")
    sys.exit(-1)

if output_file == "":
    print("An MS Excel based output file (i.e., .xlsx) must be specified.")
    sys.exit(-2)

if cf_api_endpoint == "":
    print("A Cloud Foundry API endpoint must be specified.")
    sys.exit(-3)

start(o_name, s_names, output_file, cf_api_endpoint)
