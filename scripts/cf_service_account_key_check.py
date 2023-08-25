######################################################################################################
# Background docs: https://cloud.gov/docs/compliance/auditing-activity/.
# There is a ton of duplicated code in here that I plan to refactor. Basically, v3 of the API
# seems much more consistent than v2 which will allow us to consolidate code sections.
######################################################################################################

import subprocess
import sys
import getopt
import os
import json
import hashlib
import re
from datetime import datetime, timezone, timedelta
import dateutil.parser
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

start_path = os.getcwd()
data_folder_path = start_path
cf_guid_dict = dict()
per_page = 5000
end_point_version = '/v3/'

######################################################################################################
# Generic command line functions.
######################################################################################################

def get_cf_response(url):
    if os.name == 'nt':
        # Need to escape the URL ampersands when running on Windows
        url = url.replace('&', '^&')
        p = subprocess.Popen(['cf.exe', 'curl', url],
                             stdout=subprocess.PIPE, shell=True)
    else:
        p = subprocess.Popen(['cf', 'curl', url], stdout=subprocess.PIPE)

    if p.returncode == 1:
        print("Can't find the CF CLI executable, is it installed?")
        sys.exit(1)

    output = p.communicate()[0]
    data = json.loads(output.decode('utf-8'))
    return data

def run_cmd_suppress_output(cmd, args):
    subprocess.call([cmd, args], stdout=open(
        os.devnull, "w"), stderr=subprocess.STDOUT)
    return


def run_cli_cmd(cmd, filename):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    f = open(filename + ".txt", "w")
    print(output.decode(), file=f)
    f.close()
    return


def run_api_cmd(cmd, filename):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    f = open(filename + ".json", "w")
    print(output.decode(), file=f)
    f.close()
    return



def run_api_rtn_json(url):
    all_data = {'resources': []}
    next_page = True

    while next_page:
        print(url)
        data = get_cf_response(url)
        for resource in data['resources']:
            all_data['resources'].append(resource)

        try:
            # if this element exists, a subsequent page exists so get the new URL and continue iteration
            url = data['pagination']['next']['href']
            index = url.find(end_point_version)
            url = url[index:]
        except:
            # break out of loop because this was the last page
            next_page = False

    return all_data


def run_api_cmd_rtn_json(cmd):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    return output.decode()

def run_api_cmd_and_hash_output_txt(cmd, hash_statement, json_attrib, filename):
    p1 = subprocess.Popen(cmd, stdout=subprocess.PIPE)
    output = p1.communicate()[0]
    json_data = json.loads(output)

    text = ""
    if "CF-NotAuthorized" in output.decode():
        text = "Not authorized to perform action around this action."
        hash_statement = text
        print(text + " " + '/'.join(cmd))
    elif json_attrib != None and json_attrib not in output.decode():
        text = "There is no " + json_attrib + " in the results of this action."
        hash_statement = text
    elif json_attrib == None:
        text = json_data
    else:
        text = json_data[json_attrib]

    hash_object = hashlib.md5(json.dumps(text).encode('utf-8'))
    f = open(filename + ".txt", "w")
    print(hash_statement + ": " + hash_object.hexdigest(), file=f)
    f.close()
    return


def versiontuple(v):
    return tuple(map(int, (v.split("."))))


def remove_prefix(s, prefix):
    return s[len(prefix):] if s.startswith(prefix) else s


def GetKey(val):
    for key, value in cf_guid_dict.items():
        if val == value:
            return key
    return None

######################################################################################################
######################################################################################################

def build_service_key_audit_header_row(b_audit_sheet):
    HeaderColorFill = PatternFill(fgColor="C0C0C0", fill_type="solid")
       
    b_audit_sheet["A1"] = "Organization Name"
    b_audit_sheet["A1"].fill = HeaderColorFill
    b_audit_sheet["A1"].font = Font(bold=True)

    b_audit_sheet["B1"] = "Space Name"
    b_audit_sheet["B1"].fill = HeaderColorFill
    b_audit_sheet["B1"].font = Font(bold=True)

    b_audit_sheet["C1"] = "Binding Type"
    b_audit_sheet["C1"].fill = HeaderColorFill
    b_audit_sheet["C1"].font = Font(bold=True)

    b_audit_sheet["D1"] = "Key Name"
    b_audit_sheet["D1"].fill = HeaderColorFill
    b_audit_sheet["D1"].font = Font(bold=True)

    b_audit_sheet["E1"] = "Created On"
    b_audit_sheet["E1"].fill = HeaderColorFill
    b_audit_sheet["E1"].font = Font(bold=True)

    b_audit_sheet["F1"] = "Updated On"
    b_audit_sheet["F1"].fill = HeaderColorFill
    b_audit_sheet["F1"].font = Font(bold=True)

    b_audit_sheet["G1"] = "Expires At"
    b_audit_sheet["G1"].fill = HeaderColorFill
    b_audit_sheet["G1"].font = Font(bold=True)

    b_audit_sheet["H1"] = "Expires (In days)"
    b_audit_sheet["H1"].fill = HeaderColorFill
    b_audit_sheet["H1"].font = Font(bold=True)

    b_audit_sheet["I1"] = "Service Instance Name"
    b_audit_sheet["I1"].fill = HeaderColorFill
    b_audit_sheet["I1"].font = Font(bold=True)

    b_audit_sheet["J1"] = "Service Plan Name"
    b_audit_sheet["J1"].fill = HeaderColorFill
    b_audit_sheet["J1"].font = Font(bold=True)

    b_audit_sheet["K1"] = "Service Plan Description"
    b_audit_sheet["K1"].fill = HeaderColorFill
    b_audit_sheet["K1"].font = Font(bold=True)

    b_audit_sheet["L1"] = "Organization GUID"
    b_audit_sheet["L1"].fill = HeaderColorFill
    b_audit_sheet["L1"].font = Font(bold=True)

    b_audit_sheet["M1"] = "Space GUID"
    b_audit_sheet["M1"].fill = HeaderColorFill
    b_audit_sheet["M1"].font = Font(bold=True)

    return


######################################################################################################
######################################################################################################

def service_key_audit(org_name, space_names, binding_type_filter, b_audit_sheet, cf_api_endpoint):
    print("==================================================================")
    print("Auditing service key information...")
    print("==================================================================")
    
    data_json = run_api_rtn_json("/v3/service_credential_bindings?per_page=" + str(per_page) +"")
    new_data_obj = {}
    new_data_obj["resources"] = []

    
    ExcelRowRef = 1
    for d in data_json["resources"]:
        print(ExcelRowRef)
        ExcelRowRef += 1
        
        service_instance_json = get_cf_response(remove_prefix(d["links"]["service_instance"]["href"], cf_api_endpoint))
        service_details_json = get_cf_response(remove_prefix(d["links"]["details"]["href"], cf_api_endpoint))
        
        service_plan = {}
        if (service_instance_json.get("links", {}).get("service_plan") is not None):
            service_plan_json = get_cf_response(remove_prefix(service_instance_json["links"]["service_plan"]["href"], cf_api_endpoint))
            org_space_info_lkup = cf_guid_dict.get(service_instance_json["relationships"]["space"]["data"]["guid"], {})

        if((binding_type_filter is not None and binding_type_filter == d["type"].lower()) or (binding_type_filter is None)):
            new_data_obj_entry = {}
            new_data_obj_entry["org_name"] = org_space_info_lkup["org_name"]
            new_data_obj_entry["space_name"] = org_space_info_lkup["space_name"]
            new_data_obj_entry["type"] = d["type"]
            new_data_obj_entry["name"] = d["name"]
            temp_datetime = dateutil.parser.parse(d["created_at"])
            #temp_datetime = temp_datetime.strftime("%Y-%m-%d %H:%M:%S")
            new_data_obj_entry["created_at"] = temp_datetime.strftime("%Y-%m-%d %H:%M:%S")
            temp_datetime = dateutil.parser.parse(d["updated_at"])
            #temp_datetime = temp_datetime.strftime("%Y-%m-%d %H:%M:%S")
            new_data_obj_entry["updated_at"] = temp_datetime.strftime("%Y-%m-%d %H:%M:%S")
            temp_datetime = dateutil.parser.parse(d["updated_at"])+timedelta(90)
            #temp_datetime = temp_datetime.strftime("%Y-%m-%d %H:%M:%S")
            new_data_obj_entry["expires_at"] = temp_datetime.strftime("%Y-%m-%d %H:%M:%S")
            delta = temp_datetime.replace(tzinfo=None) - datetime.now()
            new_data_obj_entry["expires_in_num_of_days"] = delta.days
            new_data_obj_entry["service_instance_name"] = service_instance_json["name"]
            if (service_plan is not None):
                new_data_obj_entry["service_plan_name"] = service_plan_json["name"]
                new_data_obj_entry["service_plan_description"] = service_plan_json["description"]
            else:
                new_data_obj_entry["service_plan_name"] = "Not applicable"
                new_data_obj_entry["service_plan_description"] = "Not applicable"
            new_data_obj_entry["org_guid"] = org_space_info_lkup["org_guid"]
            new_data_obj_entry["space_guid"] = org_space_info_lkup["space_guid"]
            new_data_obj["resources"].append(new_data_obj_entry)

    # Let's face it openpyxl doesn't really sort and pandas strips formatting...
    new_data_obj["resources"].sort(key=lambda x: (x['org_name'], x['space_name'], x['type'], x['name']))

    ExcelRowRef = 1
    for d in new_data_obj["resources"]:
       ExcelRowRef += 1
       b_audit_sheet["A" + str(ExcelRowRef)] = d["org_name"]
       b_audit_sheet["B" + str(ExcelRowRef)] = d["space_name"]
       b_audit_sheet["C" + str(ExcelRowRef)] = d["type"]
       b_audit_sheet["D" + str(ExcelRowRef)] = d["name"]
       b_audit_sheet["E" + str(ExcelRowRef)] = d["created_at"]
       b_audit_sheet["F" + str(ExcelRowRef)] = d["updated_at"]
       if d["updated_at"] is not None and d["updated_at"] != "":
            temp_datetime = dateutil.parser.parse(d["updated_at"])
            if temp_datetime.replace(tzinfo=None) < datetime.now()-timedelta(90):
                #b_audit_sheet["F" + str(ExcelRowRef)].font = Font(color='FF0000', bold=True)
                b_audit_sheet["G" + str(ExcelRowRef)].font = Font(color='FF0000', bold=True)
                b_audit_sheet["H" + str(ExcelRowRef)].font = Font(color='FF0000', bold=True)

       b_audit_sheet["G" + str(ExcelRowRef)] = d["expires_at"]
       b_audit_sheet["H" + str(ExcelRowRef)] = d["expires_in_num_of_days"]
       b_audit_sheet["I" + str(ExcelRowRef)] = d["service_instance_name"]
       b_audit_sheet["J" + str(ExcelRowRef)] = d["service_plan_name"]
       b_audit_sheet["K" + str(ExcelRowRef)] = d["service_plan_description"]
       b_audit_sheet["L" + str(ExcelRowRef)] = d["org_guid"]
       b_audit_sheet["M" + str(ExcelRowRef)] = d["space_guid"]
    
    # Let's attempt to make the column widths more reasonable by default.
    for col in b_audit_sheet.columns:
        max_lenght = 0
        col_name = re.findall('\w\d', str(col[0]))
        col_name = col_name[0]
        col_name = re.findall('\w', str(col_name))[0]
        for cell in col:
            try:
                if len(str(cell.value)) > max_lenght:
                    max_lenght = len(cell.value)
            except:
                pass
        adjusted_width = (max_lenght+2)
        b_audit_sheet.column_dimensions[col_name].width = adjusted_width

    return


######################################################################################################
######################################################################################################

def obtain_orgs_spaces(org_name, space_names):

    orgs_json = run_api_cmd_rtn_json(
        ["cf", "curl", "/v3/organizations?order_by=name&per_page=" + str(per_page) + ""])
    data = json.loads(orgs_json)

    for d in data["resources"]:
        spaces_json = run_api_cmd_rtn_json(
            ["cf", "curl", "/v3/spaces?order_by=name&per_page=" + str(per_page) + "&organization_guids="+d["guid"]])
        spaces_data = json.loads(spaces_json)
        for s in spaces_data["resources"]:
            cf_guid_dict.update({s["guid"]: {"org_guid": d["guid"], "org_name":d["name"], "space_guid": s["guid"], "space_name":s["name"]}})

    org_query = ""
    org_guid = None
    if org_name != "":
        print(cf_guid_dict)
        org_guid = GetKey(org_name)
        if org_guid == None:
            print("Could not find organization GUID for ORG NAME: " + org_name)
            sys.exit(3)
        else:
            org_query = "&organization_guids=" + org_guid

    space_query = ""
    space_guids = []
    if space_names != "":
        for value in space_names.split(","):
            space_guid = GetKey(org_name + "-" + value)
            if space_guid == None:
                print("Could not find space GUID for SPACE NAME: " + value)
                sys.exit(4)
            else:
                space_guids.append(space_guid)
        space_query = "&space_guids=" + ",".join(space_guids)

    return

def start(org_name, space_names, binding_type_filter, output_file, cf_api_endpoint):

    obtain_orgs_spaces(org_name, space_names)

    Excelworkbook = Workbook()

    b_audit_sheet = Excelworkbook.active
    b_audit_sheet.title = "Service Account Key Audit"
    build_service_key_audit_header_row(b_audit_sheet)
    service_key_audit(org_name, space_names, binding_type_filter, b_audit_sheet, cf_api_endpoint)
    Excelworkbook.save(filename=output_file)
    
    return
######################################################################################################
######################################################################################################


full_cmd_arguments = sys.argv
argument_list = full_cmd_arguments[1:]
short_options = "o:s:b:f:c:"
long_options = ["organization=", "space=", "binding_type=", "report_file=", "cf_api_endpoint="]
output_file = ""
cf_api_endpoint = ""

try:
    arguments, values = getopt.getopt(
        argument_list, short_options, long_options)
except getopt.error as err:
    print(str(err))
    sys.exit(2)

o_name = ""
s_names = ""
binding_type_filter = None
for current_argument, current_value in arguments:
    if current_argument in ("-o", "--organization"):
        o_name = current_value
    if current_argument in ("-s", "--space"):
        s_names = current_value
    if current_argument in ("-b", "--binding_type"):
        binding_type_filter = current_value
    if current_argument in ("-f", "--report_file"):
        output_file = current_value
    if current_argument in ("-c", "--cf_api_endpoint"):
        cf_api_endpoint = current_value

if o_name == "" and s_names != "":
    print("Cannot specify spaces without specifying a organization.")
    sys.exit(-1)

if output_file == "":
    print("An MS Excel based output file (i.e., .xlsx) must be specified.")
    sys.exit(-2)

if cf_api_endpoint == "":
    print("A Cloud Foundry API endpoint must be specified.")
    sys.exit(-3)

start(o_name, s_names, binding_type_filter, output_file, cf_api_endpoint)
